import os

from django.contrib.sites.shortcuts import get_current_site
from django.urls import reverse
from django.utils.encoding import smart_bytes, smart_str, DjangoUnicodeDecodeError
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from rest_framework.generics import UpdateAPIView
from .util import Util
from .serializers import KaryaIlmiahSerializer, ReviewSerializer, UserSerializer, ResetPasswordEmailRequestSerializer, SetNewPasswordSerializer
from django.http import JsonResponse, HttpResponse, HttpResponsePermanentRedirect, HttpResponseRedirect, response
from django.forms.models import model_to_dict
from rest_framework import status, generics
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.decorators import api_view, permission_classes
from .models import KaryaIlmiah, Review, User
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.contrib.sites.models import Site
from django.core.mail import send_mail
from django.conf import settings

import logging
from io import BytesIO
import xlsxwriter
import openpyxl
from decimal import Decimal

from allauth.socialaccount.providers.google.views import GoogleOAuth2Adapter

logger = logging.getLogger(__name__)
user_does_not_exist_msg = {'message': 'The user does not exist'}

class CustomRedirect(HttpResponsePermanentRedirect):
    allowed_schemes = [os.environ.get('APP_SCHEME'), 'http', 'https']


# Register API
# Will create a new user in database if valid
class Register(APIView):
    def post(self, request):
        serializer = UserSerializer(data=request.data)
        if serializer.is_valid():
            account = serializer.save()
            if account:
                return Response(status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ViewUserData(APIView):
    permission_classes = [IsAuthenticated]

    # Fetches the data of the user who is currently logged in
    def get(self, request):
        username = request.user.username
        user = User.objects.filter(username=username).first()
        serializer = UserSerializer(user)

        return Response(serializer.data)

    # Fetches the data of a user with a certain username
    def post(self, request):
        user_data = get_user_data(request)

        username = request.data['username']
        try:
            user = User.objects.filter(username=username).first()
        except User.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND) 
        serializer = UserSerializer(user)

        return Response(serializer.data, status=status.HTTP_200_OK)


# Fetches the data of the user who is currently logged in
class IsUserExist(APIView):

    def post(self, request):
        req_data = request.data

        username = req_data.get('username')
        user = User.objects.filter(username=username).first()
        print(user.email)
        if user is None:
            return Response(status=404, data="username not found")

        return Response(status=201, data="bruh")


    def get(self, request):
        username = request.user.username
        user = User.objects.filter(username=username).first()
        serializer = UserSerializer(user)

        return Response(serializer.data)

# Fetches Karils associated with the creater User based on username
class GetLinkedKarils(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        logger.info("Checking for linked karils...")
        requested_username = request.user.username
        
        karils = KaryaIlmiahSerializer(KaryaIlmiah.objects.filter(pemilik=requested_username), many=True)
        
        if len(karils.data) != 0:
            return Response(karils.data, status=status.HTTP_200_OK)
        else:
            # No matching karils
            return Response(status=status.HTTP_204_NO_CONTENT)

# Will return all user data for the given email
# only succeeds if the authenticated user's email
# is the one being queried
# For use with Google auth, to check for similar emails
class GetLinkedUsers(APIView):
    def get(self, request):
        logger.info("Checking for linked users...")
        requested_email = request.query_params['email']

        matches = User.objects.filter(email=requested_email)
        usernames = []
        for user in matches:
            usernames.append(user.username)
        logger.info(f"{len(usernames)} Matches found!")

        if len(usernames) != 0:
            response = {}
            response['usernames'] =  usernames
            return Response(response, status=status.HTTP_200_OK)
        else:
            # No matching usernames
            return Response(status=status.HTTP_204_NO_CONTENT)

# Get user data
# For use with getting user roles for authorization and others.
def get_user_data(request):
    username = request.user.username
    user = User.objects.filter(username=username).first()
    serializer = UserSerializer(user)
    return serializer.data

# User Management API
# Handles Admin/SDMPT management of user accounts
class ManageUsers(APIView):
    permission_classes = [IsAuthenticated]
    forbidden_role_msg = {'message': 'You must be an Admin or SDM PT to perform this action.'}

    # Fetches all user data
    def get(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']
        
        if ( user_role == "Admin" or user_role == "SDM PT" ):
            user_list = User.objects.all().order_by("date_joined").reverse()

            serializer = UserSerializer(user_list, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else: return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED)


    # Create new dosen
    # Same as registration..
    # but done by an authenticated admin or SDMPT.
    def post(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']
        
        if ( user_role == "Admin" or user_role == "SDM PT" ):
            register_view = Register()
            return Register.post(register_view, request)
        else: return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED)

    # Gets the user data with a certain username --> Edits according to the request
    # Username in request body
    # Accessible for Admins, SDM PT, and users who want to edit their own account.
    def put(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']

        if ( user_role == "Admin" or user_role == "SDM PT" or user_data['username'] == request.data['username']):
            try:
                user = User.objects.get(username=request.data['username'])
            except User.DoesNotExist:
                return Response(user_does_not_exist_msg, status=status.HTTP_404_NOT_FOUND)
            user = User.objects.get(username=request.data['username'])

            serializer = UserSerializer(user, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else: return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED)

    # Fetches data with a certain username, then deletes it.
    # Username in request body
    def delete(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']

        if ( user_role == "Admin" or user_role == "SDM PT" or user_data['username'] != request.data['username']):
            try:
                user = User.objects.get(username=request.data['username'])
            except User.DoesNotExist: 
                return Response(user_does_not_exist_msg, status=status.HTTP_404_NOT_FOUND) 
            user.delete()
            return Response({request.data['username'] + ' was deleted successfully!'}, status=status.HTTP_200_OK)
        else: return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED)

# Admin is able to approve user
class ApproveUsers(APIView):
    permission_classes = [IsAuthenticated]
    forbidden_role_msg = {'message': 'You must be an Admin or SDM PT to perform this action.'}
    def post(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']
        if ( user_role == "Admin" or user_role == "SDM PT" or user_data['username'] == request.data['username']):
            try:
                user = User.objects.get(username=request.data['username'])
            except User.DoesNotExist:
                return Response(user_does_not_exist_msg, status=status.HTTP_404_NOT_FOUND)
            user = User.objects.get(username=request.data['username'])
            data = {'approved' : True}
            serializer = UserSerializer(user, data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else: return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED)



# Reviewer management endpoint
# For use with Stage 2 review form creation
class ManageReviewers(APIView):
    permission_classes = [IsAuthenticated]
    forbidden_role_msg = {'message': 'You must be an Admin or SDM PT to perform this action.'}
    
    # Reviewer's position cannot be higher than reviewee
    positions = ['Asisten Ahli', 'Lektor', 'Lektor Kepala', 'Guru Besar/Professor']
    position_exclusions = {
        positions[0]: [],
        positions[1]: [positions[0]],
        positions[2]: positions[:2],
        positions[3]: positions[:3]
    }

    
    # Fetches all reviewers
    # Reviewers must be positioned equal to or higher than selected promotion rank.
    # e.g. dosen wants to be promoted to Lektor, reviewer cannot be Asisten Ahli.
    # For use with reviewer dropdown menu when assigning reviewers
    # request.data = {
    #     'position': selected promotion rank stated in review form
    # } 
    def post(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']

        if ( user_role == "Admin" or user_role == "SDM PT" ):
            # Selected promotion rank
            selected_role = request.data['position']
            user_list = (
                User.objects
                .filter(role='Reviewer')
                .exclude(position__in=self.position_exclusions[selected_role])
                .order_by("date_joined")
                .reverse()
            )
            serializer = UserSerializer(user_list, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else: return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED)

class AssignReviewer(APIView):
    permission_classes = [IsAuthenticated]
    forbidden_role_msg = {'message': 'You are not authorized to modify this review form.'}
    
    # Assign reviewer by username to certain karil
    # NOTE: User MUST assign more than 1 reviewer (see PBI 14)
    # request.data = {
    #   'reviewers': list of reviewer usernames,
    #   'karil_id': id of karil to be assigned
    # }
    def post(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']

        if ( user_role == "Admin" or user_role == "SDM PT" ):
            if len(request.data['reviewers']) < 2:
                return Response({'message': 'You must assign at least 2 reviewers!'}, status=status.HTTP_400_BAD_REQUEST)
            karil = KaryaIlmiah.objects.filter(karil_id=request.data['karil_id']).first()

            # Edit reviewers field in karil
            reviewers = User.objects.filter(username__in=request.data['reviewers'])
            serializer = KaryaIlmiahSerializer(karil, data={'reviewers': reviewers, 'status': 'Not Reviewed Yet'}, partial=True)
            if serializer.is_valid():
                reviewform = serializer.save()
                if reviewform:
                    return Response(serializer.data, status=status.HTTP_200_OK)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            else: return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(status=status.HTTP_401_UNAUTHORIZED)

class GetSpecificReviewForm(APIView):
    def post(self, request):
        try:
            karil_list = KaryaIlmiah.objects.filter(karil_id=request.data['karil_id']).first()
            serializer = KaryaIlmiahSerializer(karil_list)
        except KaryaIlmiah.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)

        return Response(serializer.data, status=status.HTTP_200_OK)

# Review form management endpoint
# For Stage 1 and Stage 2 review form creation
# NOTE: This is NOT for reviews! Only for review forms, which are basically karil entries.
# NOTE: This is also not for reviewers, see ManageReviewers and AssignReviewer
class ManageReviewForm(APIView):
    permission_classes = [IsAuthenticated]
    forbidden_role_msg = {'message': 'You are not authorized to modify this review form.'}
    serializer_class = KaryaIlmiahSerializer

    # Passes request data to serializer
    # Works just like register API
    # Creates Stage 1 review form 
    def post(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']

        # Reviewers are not allowed to create review forms!
        if (user_role == "Reviewer"):
            return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED)
        
        data = request.data
        try:
            data['pemilik'] = User.objects.filter(full_name=data['pemilik']).first().username
        except (User.DoesNotExist, AttributeError) :
            return Response({'This author does not exist!'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = KaryaIlmiahSerializer(data = request.data)
        if serializer.is_valid():
            review = serializer.save()
            if review:
                return Response({request.data['judul'] + ' was queued for review succesfully!'}, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else: return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Displays ALL submitted karils
    # Used for debugging
    # Can be deleted if unneeded
    def get(self, request):
        karil_list = KaryaIlmiah.objects.all()
        serializer = KaryaIlmiahSerializer(karil_list, many=True)

        return Response(serializer.data, status=status.HTTP_200_OK)

    # Updates Stage 1 review form into stage 2
    def put(self, request):
        """
        Stage 2 Review Form
        Admin and SDMPT can edit the review form and assign reviewers

        :param request.data['review']: existing stage 1 review form
        :return: updated stage 1 -> stage 2 review form
        """

        user_data = get_user_data(request)
        user_role = user_data['role']

        if ( user_role == "Admin" or user_role == "SDM PT" ):
            karil = None
            try:
                karil = KaryaIlmiah.objects.get(karil_id=request.data['karil_id'])
                if (karil.status != 'In Review'):   # Status cannot be 'In Review'
                    karil.status = 'Not Reviewed Yet'   ## Change status to Not Reviewed Yet after edit is called
                else: return Response({'message': 'You may not edit a paper that is in review!'}, status=status.HTTP_401_UNAUTHORIZED)

            except KaryaIlmiah.DoesNotExist:
                return Response({'message': 'This review form does not exist!'}, status=status.HTTP_404_NOT_FOUND)

            serializer = KaryaIlmiahSerializer(karil, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)

            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else: return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED)

    # Deletes karil with a requested karil_id
    # request_data = {
    #   karil_id: id of the requested karil
    # }
    def delete(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']
        try:
            karil = KaryaIlmiah.objects.get(karil_id = request.data['karil_id'])
        except KaryaIlmiah.DoesNotExist: 
            return Response({'message': 'The paper you are trying to delete does not exist'}, status=status.HTTP_404_NOT_FOUND) 
        karil_data = KaryaIlmiahSerializer(karil).data

        # Checks if a dosen is trying to delete their own karil
        if ((user_data['username'] == karil_data['pemilik'] and user_role == "Dosen") or user_role == "Admin"):
            karil.delete()
            return Response({request.data['judul'] + ' was deleted successfully!'}, status=status.HTTP_200_OK)
        else: return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED)

# Class to get reviews based on roles
class ManageKaril(APIView):
    permission_classes = [IsAuthenticated]
    forbidden_role_msg = {'message': 'You are not authorized to perform this action!'}

    def get(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']
        print("getting karils...")

        if (user_role == "Admin"):
            print("I'm an admin")
            try:
                karil_list = KaryaIlmiah.objects.all()
                serializer = KaryaIlmiahSerializer(karil_list, many=True)

            except KaryaIlmiah.DoesNotExist:
                return Response({'message': 'Papers not found!'}, status=status.HTTP_404_NOT_FOUND)
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        elif (user_role == "Reviewer"):
            print("I'm a reviewer")
            try:
                karil_list = KaryaIlmiah.objects.filter(reviewers__username = user_data['username'])
                serializer = KaryaIlmiahSerializer(karil_list, many=True)
            except KaryaIlmiah.DoesNotExist:
                return Response({'message': 'No papers are assigned to this reviewer yet! '}, status=status.HTTP_404_NOT_FOUND)
            return Response(serializer.data, status=status.HTTP_200_OK)
        else: return Response(self.forbidden_role_msg, status=status.HTTP_401_UNAUTHORIZED) 

# Review management endpoint
# NOTE: For reviews made by reviewers
class ManageKarilReview(APIView):
    permission_classes = [IsAuthenticated]

    # Create new review
    # request_data = {
    #   karil_id: id of karil to be reviewed
    #   reviewer: username of the current reviewer/admin,
    #   + any other required fields for Review (see Review model or serializer)
    # }
    def post(self, request):
        user_data = get_user_data(request)
        user_role = user_data['role']

        # Reviewers and Admins can review. Also checks if they are reviewing on their own behalf
        if ((user_role == 'Reviewer' or user_role == 'Admin') and request.data['reviewer'] == user_data['username']):
            try:
                karil = KaryaIlmiah.objects.get(karil_id = request.data['karil_id'])
            except KaryaIlmiah.DoesNotExist:
                return Response({'message': 'The paper you are trying to review does not exist!'}, status=status.HTTP_404_NOT_FOUND)
            
            serializer = ReviewSerializer(data=request.data)
            if serializer.is_valid():
                new = serializer.save()
                
                # Update karil reviews
                review_id = new.review_id
                print(review_id)
                karil_reviews = KaryaIlmiahSerializer(karil).data['reviews']
                karil_reviews.append(review_id)
                karil_serializer = KaryaIlmiahSerializer(karil, data={'reviews': karil_reviews}, partial=True)
                karil_serializer.is_valid(raise_exception=True)
                karil_serializer.save()

                return Response({'message': 'Review has successfully been created!'}, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else: return Response(status=status.HTTP_401_UNAUTHORIZED)

    # Get review by id in request params (see url)
    def get(self, request):
        review_id = request.query_params.get('id')
        try:
            reviews = Review.objects.all().filter(review_id=review_id)
        except Review.DoesNotExist:
            return Response({'message': 'This review does not exist!'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ReviewSerializer(reviews, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class GetLinkedReviews(APIView):
    permission_classes = [IsAuthenticated]
    # Get review by id in request params (see url)
    def get(self, request):
        karil_id = request.query_params.get('id')
        try:
            karil = KaryaIlmiah.objects.get(karil_id = karil_id)
        except KaryaIlmiah.DoesNotExist:
            return Response({'message': 'This karil does not exist!'}, status=status.HTTP_404_NOT_FOUND)
        
        review_id_list = KaryaIlmiahSerializer(karil).data['reviews']
        try:
            reviews = Review.objects.all().filter(review_id__in=review_id_list)
        except Review.DoesNotExist:
            return Response({'message': 'No Reviews linked to this karil!'}, status=status.HTTP_404_NOT_FOUND)

        serializer = ReviewSerializer(reviews, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class GetAssignedKarils(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user_data = get_user_data(request)
        try:
            karils = KaryaIlmiah.objects.all().filter(reviewers__username = user_data['username'])
        except KaryaIlmiah.DoesNotExist:
            return Response({'message': 'This review does not exist!'}, status=status.HTTP_404_NOT_FOUND)
        serializer = KaryaIlmiahSerializer(karils, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class RequestPasswordResetEmail(generics.GenericAPIView):
    serializer_class = ResetPasswordEmailRequestSerializer
    def post(self, request):
        serializer = ResetPasswordEmailRequestSerializer(data=request.data)
        if serializer.is_valid():
            username = serializer.data['username']
            if User.objects.filter(username=username).exists():
                user = User.objects.get(username=username)
                if user is None:
                    return Response({'failed': "user not found"}, status=404)
                uidb64 = urlsafe_base64_encode(smart_bytes(user.pk))
                id = smart_str(urlsafe_base64_decode(uidb64))
         
                token = PasswordResetTokenGenerator().make_token(user)
                current_site = request.get_host()
                relative_link = reverse('main:password-reset-confirm',
                                       kwargs={'uidb64': uidb64, 'token': token})
                absurl = 'http://' + current_site + relative_link
                subject = 'Reset your password'
                email_from = settings.EMAIL_HOST_USER
                list_email_to = [user.email,]
                email_body = 'Hello, \n Use link below to reset your password  \n' + \
                             absurl 
                
                data = {'email_body': email_body, 'to_email': user.email,
                        'email_subject': 'Reset your passsword'}
                send_mail(subject, email_body, email_from, list_email_to)
         
                print(email_from)
                print(user.email)
     
                return Response({'success': 'We have sent you a link to reset your password', 'data': absurl},
                                status=status.HTTP_200_OK)
            else:
                return Response({'failed': "user not found"}, status=404)


class PasswordTokenCheckAPI(generics.GenericAPIView):
    serializer_class = SetNewPasswordSerializer

    def get(self, request, uidb64, token):
        current_site = settings.FRONTEND_REDIRECT
        try:
            id = smart_str(urlsafe_base64_decode(uidb64))
            user = User.objects.filter(username=id).first()
            if not PasswordResetTokenGenerator().check_token(user, token):
                print("invalid token")
                return HttpResponseRedirect(current_site + "/token-error")

            return HttpResponseRedirect( current_site + "/reset-password/" + token + "/" + uidb64)


        except DjangoUnicodeDecodeError as identifier:
            try:
                if not PasswordResetTokenGenerator().check_token(user):
                    return HttpResponseRedirect(current_site + "/token-error")


            except UnboundLocalError:
                return HttpResponseRedirect(current_site + "/token-error")


class SetNewPasswordAPIView(generics.GenericAPIView):
    serializer_class = SetNewPasswordSerializer

    def patch(self, request):
        print(request.data)
        serializer = self.serializer_class(data=request.data)
        serializer.is_valid(raise_exception=True)
        return Response({'success': True, 'message': 'Password reset success'}, status=status.HTTP_200_OK)

   

# Test view for user authentication
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def ping_auth(request):
    return JsonResponse({'message': 'You are logged in!'})


# General ping test
def ping(request):
    if request.method == "GET":
        return JsonResponse({'foo': 'bar'})


# Handles downloading of Karil Info
# Needs karil_id in request body
# Will download an .xlsx file
class DownloadKaril(APIView):
    def post(self, request):
        logger.info("start of function")

        row_num = 0
        col_num = 0

        logger.info("declaring column names...")
        columns = ['Nama Penulis', 'Judul Karya Ilmiah', 'Data Jurnal', 'Link Asli Jurnal', 'Link Repository', 
                    'Link Indexer', 'Link Check Similarity', 'Link Bukti Korespondensi', 
                    'Peng-Index', 'Kategori Karya Ilmiah', ]

        try:
            # Get requested karil data
            logger.info("getting karil info...")
            karil = KaryaIlmiah.objects.get(karil_id = request.data['karil_id'])
            
        except KaryaIlmiah.DoesNotExist:
            logger.warn("none found...")
            return Response({'message': 'This review does not exist!'}, status=status.HTTP_404_NOT_FOUND)
        
        wb_title=karil.judul+'.xlsx'

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet1 = workbook.add_worksheet()

        bold = workbook.add_format({'bold': True, 'right':2})

        worksheet1.set_column(col_num, col_num+1, 25)

        for row_num in range(len(columns)):
            worksheet1.write(row_num, col_num, columns[row_num], bold)

        col_num += 1

        logger.info("writing karil info...")
        worksheet1.write(0, col_num, karil.pemilik.username) # format for ws.write is row_num, col_num, data, font styling
        worksheet1.write(1, col_num, karil.judul)
        worksheet1.write(2, col_num, karil.journal_data)
        worksheet1.write_url(3, col_num, karil.link_origin)
        worksheet1.write_url(4, col_num, karil.link_repo)
        worksheet1.write_url(5, col_num, karil.link_indexer)
        worksheet1.write_url(6, col_num, karil.link_simcheck)
        worksheet1.write_url(7, col_num, karil.link_correspondence)
        worksheet1.write(8, col_num, karil.indexer)
        worksheet1.write(9, col_num, karil.category)

        logger.info("saving workbook...")
        workbook.close()

        output.seek(0)
        # xlsx_data = output.getvalue()

        logger.info("serving download...")

        response = HttpResponse(output.read(), content_type = 'application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(wb_title)

        output.close()

        return response

# Handles downloading of Karil Review Form
# Needs karil_id in request body
# Will download an .xlsx file
# Super long lmao
class DownloadReviewForm(APIView):
    def post(self, request):
        logger.info("start of function")

        try:
            # Get requested karil data
            logger.info("getting karil info...")
            karil = KaryaIlmiah.objects.get(karil_id = request.data['karil_id'])
            
        except KaryaIlmiah.DoesNotExist:
            logger.warn("none found...")
            return Response({'message': 'This review does not exist!'}, status=status.HTTP_404_NOT_FOUND)
        
        wb_title=karil.judul+' Review Form.xlsx'

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet1 = workbook.add_worksheet()

        merge_format = workbook.add_format({
                        'align': 'center',
                        'valign': 'vcenter',
                        'text_wrap' : True
                        })
        
        merge_format_bordered = workbook.add_format({
                                'align': 'center',
                                'valign': 'vcenter',
                                'text_wrap' : True,
                                'top' : 1,
                                'left' : 1,
                                'right' : 1,
                                'bottom' : 1
                        })
        
        merge_format_number = workbook.add_format({
                        'align': 'center',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'bottom': 1
                        })

        merge_format_number_2 = workbook.add_format({
                        'align': 'center',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'bottom': 5
                        })

        merge_format_number_bordered = workbook.add_format({
                        'align': 'center',
                        'valign' : 'vcenter',
                        'text_wrap' : True,
                        'top' : 1,
                        'bottom' : 1,
                        'left' : 1,
                        'right' : 1
                        })

        merge_format_justified_category_1 = workbook.add_format({
                        'align': 'vjustify',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'bottom': 1,
                        'left': 1
                        })

        merge_format_justified_category_2 = workbook.add_format({
                        'align': 'vjustify',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'left': 1
                        })
        
        merge_format_justified_category_3 = workbook.add_format({
                        'align': 'vjustify',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'left': 1
                        })

        merge_format_justified_category_4 = workbook.add_format({
                        'align': 'vjustify',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'bottom': 1
                        })
        
        merge_format_justified_category_5 = workbook.add_format({
                        'align': 'vjustify',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'bottom': 5,
                        'left': 1
                        })
        
        merge_format_justified_category_6 = workbook.add_format({
                        'align': 'vjustify',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'bottom': 5,
                        'left': 1,
                        'right': 1
                        })
        
        merge_format_justified_category_bordered = workbook.add_format({
                        'align': 'vjustify',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'bottom': 1,
                        'left': 1,
                        'right': 1
                        })

        merge_format_justified_content = workbook.add_format({
                        'align': 'vjustify',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'bottom': 1,
                        'right': 5
                        })

        merge_format_justified_content_2 = workbook.add_format({
                        'align': 'vjustify',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'bottom': 5,
                        'right': 5
                        })
        
        merge_format_justified_content_3 = workbook.add_format({
                        'align': 'center',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'top': 1,
                        'bottom': 1,
                        'right': 5
                        })
        
        merge_format_title = workbook.add_format({
                        'align': 'center',
                        'valign': 'vcenter',
                        'text_wrap' : True,
                        'bottom' : 1,
                        'right' : 5
                        })

        bold = workbook.add_format({'bold': True})
        italic = workbook.add_format({'italic': True})
        bolditalic = workbook.add_format({'bold': True, 'italic': True})

        worksheet1.set_column(0, 0, 3)
        worksheet1.set_column(1, 6, 16)


        ## WRITING TITLE
        worksheet1.merge_range('A1:G1', '', merge_format_title)
        worksheet1.write_rich_string('A1', bold, 'HASIL PENILAIAN SEJAWAT SEBIDANG ATAU ', bolditalic, 'PEER REVIEW', merge_format_title)
        worksheet1.write('A2', 'A', merge_format_number)
        worksheet1.merge_range('B2:G2', '', merge_format_justified_content)
        worksheet1.write('B2', 'Identitas karya ilmiah (diisi sesuai standar referensi)')


        ## == KARIL INFO ==

        worksheet1.set_column('H:H', None, None, {'hidden': 1})
        worksheet1.write('H1', karil.karil_id, merge_format)

        worksheet1.set_row(2, 50, merge_format)
        worksheet1.write('A3', '1', merge_format_number)
        worksheet1.write('B3', 'Judul', merge_format_justified_category_1)
        worksheet1.merge_range('C3:G3', '', merge_format_justified_content)
        worksheet1.write_rich_string('C3', merge_format, ': ', karil.judul, merge_format_justified_content)

        worksheet1.set_row(3, 35, merge_format)
        worksheet1.write('A4', '2', merge_format_number)
        worksheet1.write('B4', 'Data Jurnal', merge_format_justified_category_2)
        worksheet1.merge_range('C4:G4', '', merge_format_justified_content)
        worksheet1.write_rich_string('C4', merge_format, ': ', karil.journal_data, merge_format_justified_content)

        worksheet1.set_row(4, 30, merge_format)
        worksheet1.write('A5', '3.1', merge_format_number)
        worksheet1.write('B5', 'Nama Penulis', merge_format_justified_category_3)
        worksheet1.merge_range('C5:G5', '', merge_format_justified_content)
        worksheet1.write_rich_string('C5', merge_format, ': ', karil.pemilik.username, merge_format_justified_content)

        worksheet1.set_row(5, 30, merge_format)
        worksheet1.write('A6', '3.2', merge_format_number)
        worksheet1.write('B6', 'Link Asli Jurnal', merge_format_justified_category_3)
        worksheet1.merge_range('C6:G6', '', merge_format_justified_content)
        worksheet1.write_url('C6', karil.link_origin, merge_format_justified_content)
        
        worksheet1.set_row(6, 30, merge_format)
        worksheet1.write('A7', '3.3', merge_format_number)
        worksheet1.write('B7', 'Link Repositori Jurnal', merge_format_justified_category_3)
        worksheet1.merge_range('C7:G7', '', merge_format_justified_content)
        worksheet1.write_url('C7', karil.link_repo, merge_format_justified_content)

        worksheet1.set_row(7, 30, merge_format)
        worksheet1.write('A8', '3.4', merge_format_number)
        worksheet1.write('B8', 'Link Indexer', merge_format_justified_category_3)
        worksheet1.merge_range('C8:G8', '', merge_format_justified_content)
        worksheet1.write_url('C8', karil.link_indexer, merge_format_justified_content)

        worksheet1.set_row(8, 30, merge_format)
        worksheet1.write('A9', '3.5', merge_format_number)
        worksheet1.write('B9', 'Link Check Similarity', merge_format_justified_category_3)
        worksheet1.merge_range('C9:G9', '', merge_format_justified_content)
        worksheet1.write_url('C9', karil.link_simcheck, merge_format_justified_content)

        worksheet1.set_row(9, 30, merge_format)
        worksheet1.write('A10', '3.6', merge_format_number)
        worksheet1.write('B10', 'Link Bukti Correspondence', merge_format_justified_category_4)
        worksheet1.merge_range('C10:G10', '', merge_format_justified_content)
        worksheet1.write_url('C10', karil.link_correspondence, merge_format_justified_content)

        worksheet1.set_row(10, 30, merge_format)
        worksheet1.write('A11', 'B', merge_format_number)
        worksheet1.write('B11', 'Peng-index (jika ada)', merge_format_justified_category_1)
        worksheet1.merge_range('C11:G11', '', merge_format_justified_content)
        worksheet1.write_rich_string('C11', merge_format, ': ', karil.indexer, merge_format_justified_content)


        ## == KARIL CATEGORIES (35 kinds) ==
        worksheet1.set_row(11, 30, merge_format)
        worksheet1.write('A12', 'C', merge_format_number)
        worksheet1.merge_range('B12:F12', '', merge_format_justified_content)
        worksheet1.write('B12', 'Kategori karya ilmiah, dan nilai maksimal (pilih salah satu dengan memberikan tanda P)', merge_format_justified_category_1)
        worksheet1.write('G12', 'AK', merge_format_justified_content_3)

        worksheet1.set_row(12, 18, merge_format)
        worksheet1.merge_range('A13:G13', '', merge_format_justified_content)
        worksheet1.write('A13', 'Buku', merge_format_justified_content)

        worksheet1.write('A14', '1', merge_format_number)
        worksheet1.merge_range('B14:F14', '', merge_format_justified_content)
        worksheet1.write('B14', 'Buku referensi', merge_format_justified_category_1)
        worksheet1.write('G14', '40', merge_format_justified_content_3)

        worksheet1.write('A15', '2', merge_format_number)
        worksheet1.merge_range('B15:F15', '', merge_format_justified_content)
        worksheet1.write('B15', 'Buku monograph', merge_format_justified_category_1)
        worksheet1.write('G15', '20', merge_format_justified_content_3)

        worksheet1.write('A16', '3', merge_format_number)
        worksheet1.merge_range('B16:F16', '', merge_format_justified_content)
        worksheet1.write_rich_string('B16', italic, 'Book chapter', '(internasional)', merge_format_justified_category_1)
        worksheet1.write('G16', '15', merge_format_justified_content_3)

        worksheet1.write('A17', '4', merge_format_number)
        worksheet1.merge_range('B17:F17', '', merge_format_justified_content)
        worksheet1.write_rich_string('B17', italic, 'Book chapter', '(nasional)', merge_format_justified_category_1)
        worksheet1.write('G17', '10', merge_format_justified_content_3)

        worksheet1.set_row(17, 18, merge_format)
        worksheet1.merge_range('A18:G18', '', merge_format_justified_content)
        worksheet1.write('A18', 'Jurnal', merge_format_justified_content)

        worksheet1.write('A19', '5', merge_format_number)
        worksheet1.merge_range('B19:F19', '', merge_format_justified_content)
        worksheet1.write('B19', 'Jurnal internasional bereputasi (terindeks pada database internasional bereputasi dan berfaktor dampak)', merge_format_justified_category_1)
        worksheet1.write('G19', '40', merge_format_justified_content_3)

        worksheet1.write('A20', '6', merge_format_number)
        worksheet1.merge_range('B20:F20', '', merge_format_justified_content)
        worksheet1.write('B20', 'Jurnal internasional terindeks pada basis data internasional bereputasi', merge_format_justified_category_1)
        worksheet1.write('G20', '30', merge_format_justified_content_3)

        worksheet1.write('A21', '7', merge_format_number)
        worksheet1.merge_range('B21:F21', '', merge_format_justified_content)
        worksheet1.write('B21', 'Jurnal internasional terindeks pada basis data non bereputasi', merge_format_justified_category_1)
        worksheet1.write('G21', '20', merge_format_justified_content_3)

        worksheet1.write('A22', '8', merge_format_number)
        worksheet1.merge_range('B22:F22', '', merge_format_justified_content)
        worksheet1.write('B22', 'Jurnal nasional terakreditasi Kemenristek Dikti', merge_format_justified_category_1)
        worksheet1.write('G22', '25', merge_format_justified_content_3)

        worksheet1.write('A23', '9', merge_format_number)
        worksheet1.merge_range('B23:F23', '', merge_format_justified_content)
        worksheet1.write('B23', 'Jurnal nasional terakreditasi Kemenristek Dikti peringkat 1 dan 2', merge_format_justified_category_1)
        worksheet1.write('G23', '25', merge_format_justified_content_3)

        worksheet1.set_row(23, 30, merge_format)
        worksheet1.write('A24', '10', merge_format_number)
        worksheet1.merge_range('B24:F24', '', merge_format_justified_content)
        worksheet1.write('B24', 'Jurnal nasional berbahasa Inggris atau bahasa resmi (PBB) terindeks pada basis data yang diakui Kemenristekdikti, contoh: CABI atau Index Copernicus International (ICI)', merge_format_justified_category_1)
        worksheet1.write('G24', '20', merge_format_justified_content_3)

        worksheet1.set_row(24, 30, merge_format)
        worksheet1.write('A25', '11', merge_format_number)
        worksheet1.merge_range('B25:F25', '', merge_format_justified_content)
        worksheet1.write('B25', 'Jurnal nasional berbahasa Indonesia terindeks pada basis data yang diakui Kemenristekdikti, contoh : akreditasi peringkat 5 dan 6', merge_format_justified_category_1)
        worksheet1.write('G25', '15', merge_format_justified_content_3)

        worksheet1.write('A26', '12', merge_format_number)
        worksheet1.merge_range('B26:F26', '', merge_format_justified_content)
        worksheet1.write('B26', 'Jurnal nasional', merge_format_justified_category_1)
        worksheet1.write('G26', '10', merge_format_justified_content_3)

        worksheet1.set_row(26, 30, merge_format)
        worksheet1.write('A27', '13', merge_format_number)
        worksheet1.merge_range('B27:F27', '', merge_format_justified_content)
        worksheet1.write('B27', 'Jurnal ilmiah yang ditulis dalam Bahasa Resmi PBB namun tidak memenuhi syarat syarat sebagai jurnal ilmiah internasional', merge_format_justified_category_1)
        worksheet1.write('G27', '10', merge_format_justified_content_3)

        worksheet1.set_row(27, 18, merge_format)
        worksheet1.merge_range('A28:G28', '', merge_format_justified_content)
        worksheet1.write('A28', 'Dipresentasikan secara oral dan dimuat dalam prosiding yang dipublikasikan (ber ISSN/ISBN)', merge_format_justified_content)

        worksheet1.write('A29', '14', merge_format_number)
        worksheet1.merge_range('B29:F29', '', merge_format_justified_content)
        worksheet1.write('B29', 'Internasional terindeks pada Scimagojr dan Scopus', merge_format_justified_category_1)
        worksheet1.write('G29', '30', merge_format_justified_content_3)

        worksheet1.write('A30', '15', merge_format_number)
        worksheet1.merge_range('B30:F30', '', merge_format_justified_content)
        worksheet1.write('B30', 'Internasional terindeks pada Scopus/IEEE Explore/SPIE', merge_format_justified_category_1)
        worksheet1.write('G30', '25', merge_format_justified_content_3)

        worksheet1.write('A31', '16', merge_format_number)
        worksheet1.merge_range('B31:F31', '', merge_format_justified_content)
        worksheet1.write('B31', 'Internasional', merge_format_justified_category_1)
        worksheet1.write('G31', '15', merge_format_justified_content_3)

        worksheet1.write('A32', '17', merge_format_number)
        worksheet1.merge_range('B32:F32', '', merge_format_justified_content)
        worksheet1.write('B32', 'Nasional', merge_format_justified_category_1)
        worksheet1.write('G32', '10', merge_format_justified_content_3)

        worksheet1.set_row(32, 18, merge_format)
        worksheet1.merge_range('A33:G33', '', merge_format_justified_content)
        worksheet1.write('A33', 'Disajikan dalam bentuk poster dan dimuat dalam prosiding yang dipublikasikan', merge_format_justified_content)

        worksheet1.write('A34', '18', merge_format_number)
        worksheet1.merge_range('B34:F34', '', merge_format_justified_content)
        worksheet1.write('B34', 'Internasional', merge_format_justified_category_1)
        worksheet1.write('G34', '10', merge_format_justified_content_3)

        worksheet1.write('A35', '19', merge_format_number)
        worksheet1.merge_range('B35:F35', '', merge_format_justified_content)
        worksheet1.write('B35', 'Nasional', merge_format_justified_category_1)
        worksheet1.write('G35', '5', merge_format_justified_content_3)

        worksheet1.set_row(35, 18, merge_format)
        worksheet1.merge_range('A36:G36', '', merge_format_justified_content)
        worksheet1.write('A36', 'Disajikan dalam seminar/simposium/lokakarya, tetapi tidak dimuat dalam prosiding yang dipublikasikan', merge_format_justified_content)

        worksheet1.write('A37', '20', merge_format_number)
        worksheet1.merge_range('B37:F37', '', merge_format_justified_content)
        worksheet1.write('B37', 'Internasional (bukti sertifikat)', merge_format_justified_category_1)
        worksheet1.write('G37', '5', merge_format_justified_content_3)

        worksheet1.write('A38', '21', merge_format_number)
        worksheet1.merge_range('B38:F38', '', merge_format_justified_content)
        worksheet1.write('B38', 'Nasional (bukti sertifikat)', merge_format_justified_category_1)
        worksheet1.write('G38', '3', merge_format_justified_content_3)

        worksheet1.set_row(38, 18, merge_format)
        worksheet1.merge_range('A39:G39', '', merge_format_justified_content)
        worksheet1.write('A39', 'Hasil penelitian/pemikiran yang tidak disajikan dalam seminar/simposium/lokakarya, tetapi dimuat dalam prosiding', merge_format_justified_content)

        worksheet1.write('A40', '22', merge_format_number)
        worksheet1.merge_range('B40:F40', '', merge_format_justified_content)
        worksheet1.write('B40', 'Internasional', merge_format_justified_category_1)
        worksheet1.write('G40', '10', merge_format_justified_content_3)

        worksheet1.write('A41', '23', merge_format_number)
        worksheet1.merge_range('B41:F41', '', merge_format_justified_content)
        worksheet1.write('B41', 'Nasional', merge_format_justified_category_1)
        worksheet1.write('G41', '5', merge_format_justified_content_3)

        worksheet1.set_row(41, 21, merge_format)
        worksheet1.write('A42', '24', merge_format_number)
        worksheet1.merge_range('B42:F42', '', merge_format_justified_content)
        worksheet1.write('B42', 'Hasil penelitian/pemikiran yang disajikan dalam koran/majalah populer/umum', merge_format_justified_category_1)
        worksheet1.write('G42', '1', merge_format_justified_content_3)

        worksheet1.set_row(42, 30, merge_format)
        worksheet1.write('A43', '25', merge_format_number)
        worksheet1.merge_range('B43:F43', '', merge_format_justified_content)
        worksheet1.write('B43', 'Hasil penelitian atau pemikiran atau kerjasama industri yang tidak dipublikasikan (tersimpan dalam perpustakaan) yang dilakukan secara melembaga', merge_format_justified_category_1)
        worksheet1.write('G43', '2', merge_format_justified_content_3)

        worksheet1.write('A44', '26', merge_format_number)
        worksheet1.merge_range('B44:F44', '', merge_format_justified_content)
        worksheet1.write('B44', 'Menerjemahkan/menyadur buku ilmiah, diterbitkan dan diedarkan secara nasional', merge_format_justified_category_1)
        worksheet1.write('G44', '15', merge_format_justified_content_3)

        worksheet1.write('A45', '27', merge_format_number)
        worksheet1.merge_range('B45:F45', '', merge_format_justified_content)
        worksheet1.write('B45', 'Mengedit/menyunting karya ilmiah, diterbitkan dan diedarkan secara nasional', merge_format_justified_category_1)
        worksheet1.write('G45', '10', merge_format_justified_content_3)

        worksheet1.set_row(45, 30, merge_format)
        worksheet1.merge_range('A46:G46', '', merge_format_justified_content)
        worksheet1.write('A46', 'HAKI - Membuat rancangan dan karya teknologi yang dipatenkan atau seni yang terdaftar di HAKI secara nasional atau internasional', merge_format_justified_content)

        worksheet1.write('A47', '28', merge_format_number)
        worksheet1.merge_range('B47:F47', '', merge_format_justified_content)
        worksheet1.write('B47', 'Internasional (paling sedikit diakui oleh 4 negara)', merge_format_justified_category_1)
        worksheet1.write('G47', '60', merge_format_justified_content_3)

        worksheet1.write('A48', '29', merge_format_number)
        worksheet1.merge_range('B48:F48', '', merge_format_justified_content)
        worksheet1.write('B48', 'Nasional', merge_format_justified_category_1)
        worksheet1.write('G48', '40', merge_format_justified_content_3)

        worksheet1.set_row(48, 30, merge_format)
        worksheet1.write('A49', '30', merge_format_number)
        worksheet1.merge_range('B49:F49', '', merge_format_justified_content)
        worksheet1.write('B49', 'Nasional, dalam bentuk paten sederhana yang telah memiliki sertifikat dari Direktorat Jenderal Kekayaan Intelektual, Kemenkumham', merge_format_justified_category_1)
        worksheet1.write('G49', '20', merge_format_justified_content_3)

        worksheet1.set_row(49, 30, merge_format)
        worksheet1.write('A50', '31', merge_format_number)
        worksheet1.merge_range('B50:F50', '', merge_format_justified_content)
        worksheet1.write('B50', 'Karya ciptaan desain industri, indikasi geografis yang telah memiliki sertifikat dari Direktorat Jenderal Kekayaan Intelektual, Kemenkumham (termasuk kategori ini: Buku/Modul Ajar', merge_format_justified_category_1)
        worksheet1.write('G50', '15', merge_format_justified_content_3)

        worksheet1.set_row(50, 30, merge_format)
        worksheet1.merge_range('A51:G51', '', merge_format_justified_content)
        worksheet1.write('A51', 'Membuat rancangan dan karya teknologi yang tidak dipatenkan; rancangan dan karya seni monumental yang tidak terdaftar di HAKI tetapi telah dipresentasikan pada forum yang teragenda', merge_format_justified_content)

        worksheet1.write('A52', '32', merge_format_number)
        worksheet1.merge_range('B52:F52', '', merge_format_justified_content)
        worksheet1.write('B52', 'Internasional', merge_format_justified_category_1)
        worksheet1.write('G52', '20', merge_format_justified_content_3)

        worksheet1.write('A53', '33', merge_format_number)
        worksheet1.merge_range('B53:F53', '', merge_format_justified_content)
        worksheet1.write('B53', 'Nasional', merge_format_justified_category_1)
        worksheet1.write('G53', '15', merge_format_justified_content_3)

        worksheet1.write('A54', '34', merge_format_number)
        worksheet1.merge_range('B54:F54', '', merge_format_justified_content)
        worksheet1.write('B54', 'Lokal', merge_format_justified_category_1)
        worksheet1.write('G54', '10', merge_format_justified_content_3)

        worksheet1.write('A55', '35', merge_format_number_2)
        worksheet1.merge_range('B55:F55', '', merge_format_justified_content_2)
        worksheet1.write('B55', 'Rancangan dan karya seni yang tidak terdaftar HAKI', merge_format_justified_category_5)
        worksheet1.write('G55', '', merge_format_justified_content_2)
        ## ================================================================================================================= ##

        ## == COMMENT ON LINEARITY AND PLAGIARISM == ##
        worksheet1.merge_range('A58:G58', 'Hasil penilaian validasi', merge_format_bordered)

        worksheet1.write('B59', 'Aspek', merge_format_bordered)
        worksheet1.merge_range('C59:G59', 'Uraian/komentar penilaian', merge_format_bordered)

        worksheet1.write('A60', 'D', merge_format_number_bordered)
        worksheet1.set_row(59, 60, merge_format)
        worksheet1.write_rich_string('B60', merge_format_justified_category_bordered, 'Indikasi plagiasi (lihat ', italic, 'check similarity)', merge_format_justified_category_bordered)
        worksheet1.merge_range('C60:G60', '', merge_format_justified_category_bordered)

        worksheet1.write('A61', 'E', merge_format_number_bordered)
        worksheet1.set_row(60, 60, merge_format)
        worksheet1.write('B61', 'Linearitas', merge_format_justified_category_bordered)
        worksheet1.merge_range('C61:G61', '', merge_format_justified_category_bordered)

        worksheet1.merge_range('A62:G62', '', merge_format_bordered)
        worksheet1.write_rich_string('A62', merge_format_bordered, 'Hasil penilaian ', italic, 'peer review', merge_format_bordered)

        ## == PEER REVIEW GRADING ==
        ## AUTO-GENERATED ACCORDING TO CHOSEN CATEGORY
        worksheet1.set_row(62, 60, merge_format)
        worksheet1.write('B63', 'Komponen yang dinilai', merge_format_bordered)
        worksheet1.merge_range('C63:D63', 'Komentar/ulasan peer reviewer (wajib diisi dan hindari komentar yang hanya memuat satu dua kata seperti: sangat dalam, cukup baik, sangat berkualitas)', merge_format_justified_category_bordered)
        worksheet1.write('E63', 'Nilai Maks', merge_format_bordered)
        worksheet1.merge_range('F63:G63', 'Nilai Akhir', merge_format_bordered)

        worksheet1.write('A64', 'F', merge_format_number_bordered)
        worksheet1.set_row(63, 75, merge_format)
        worksheet1.write('B64', 'Kelengkapan dan kesesuaian unsur publikasi (10%)', merge_format_justified_category_bordered)
        worksheet1.merge_range('C64:D64', '', merge_format_justified_category_bordered)

        worksheet1.write('A65', 'G', merge_format_number_bordered)
        worksheet1.set_row(64, 90, merge_format)
        worksheet1.write('B65', 'Ruang lingkup,  kedalaman pembahasan, keterbaruan (30%)', merge_format_justified_category_bordered)
        worksheet1.merge_range('C65:D65', '', merge_format_justified_category_bordered)

        worksheet1.write('A66', 'H', merge_format_number_bordered)
        worksheet1.set_row(65, 90, merge_format)
        worksheet1.write('B66', 'Kecukupan dan kemutakhiran data/informasi dan metodologi (30%)', merge_format_justified_category_bordered)
        worksheet1.merge_range('C66:D66', '', merge_format_justified_category_bordered)

        worksheet1.write('A67', 'I', merge_format_number_bordered)
        worksheet1.set_row(66, 75, merge_format)
        worksheet1.write('B67', 'Kelengkapan unsur dan kualitas penerbit, hasil dan manfaat (30%)', merge_format_justified_category_bordered)
        worksheet1.merge_range('C67:D67', '', merge_format_justified_category_bordered)

        worksheet1.write('A68', 'J', merge_format_number_bordered)
        worksheet1.set_row(67, 30, merge_format)
        worksheet1.merge_range('B68:D68', 'Jumlah Total = (100%)', merge_format_justified_category_bordered)

        ## CONDITIONAL BASED ON CHOSEN CATEGORY ##

        ### Lists of categories based on max nilai:
        list_60 = ["""HAKI - Membuat rancangan dan karya teknologi yang dipatenkan atau seni yang terdaftar di HAKI secara nasional atau internasional - 
        Internasional (paling sedikit diakui oleh 4 negara)"""]

        list_40 = ["Buku referensi", 
        "Jurnal internasional bereputasi (terindeks pada database internasional bereputasi dan berfaktor", 
        "HAKI - Membuat rancangan dan karya teknologi yang dipatenkan atau seni yang terdaftar di HAKI secara nasional atau internasional - Nasional"]

        list_30 = ["Jurnal internasional terindeks pada basis data internasional bereputasi", 
        "Dipresentasikan secara oral dan dimuat dalam prosiding yang dipublikasikan (ber ISSN/ISBN) - Internasional terindeks pada Scimagojr dan Scopus"]
        
        list_25 = ["Jurnal nasional terakreditasi Kemenristek Dikti" or "Jurnal nasional terakreditasi Kemenristek Dikti peringkat 1 dan 2", 
        "Dipresentasikan secara oral dan dimuat dalam prosiding yang dipublikasikan (ber ISSN/ISBN) - Internasional terindeks pada Scopus/IEEE Explore/SPIE"]

        list_20 = ["Buku monograph", 
        "Jurnal internasional terindeks pada basis data non bereputasi", 
        """Jurnal nasional berbahasa Inggris atau bahasa resmi (PBB) terindeks pada basis data yang diakui Kemenristekdikti, 
        contoh: CABI atau Index Copernicus International (ICI)""", 
        """HAKI - Membuat rancangan dan karya teknologi yang dipatenkan atau seni yang terdaftar di HAKI secara nasional atau internasional - 
        Nasional, dalam bentuk paten sederhana yang telah memiliki sertifikat dari Direktorat Jenderal Kekayaan Intelektual, Kemenkumham""", 
        """Membuat rancangan dan karya teknologi yang tidak dipatenkan; rancangan dan karya seni monumental yang tidak terdaftar di HAKI tetapi 
        telah dipresentasikan pada forum yang teragenda - Internasional"""]

        list_15 = ["Book chapter (internasional)", 
        "Jurnal nasional berbahasa Indonesia terindeks pada basis data yang diakui Kemenristekdikti, contoh: akreditasi peringkat 5 dan 6", 
        "Direpresentasikan secara oral dan dimuat dalam prosiding yang dipublikasikan (ber ISSN/ISBN) - Internasional", 
        """Hasil penelitian/pemikiran yang tidak disajikan dalam seminar/simposium/lokakarya, tetapi dimuat dalam prosiding - 
        Menerjemahkan/menyadur buku ilmiah, diterbitkan dan diedarkan secara nasional""", 
        """HAKI - Membuat rancangan dan karya teknologi yang dipatenkan atau seni yang terdaftar di HAKI secara nasional atau internasional - 
        Karya ciptaan desain industri, indikasi geografis yang telah memiliki sertifikat dari Direktorat Jenderal Kekayaan Intelektual, Kemenkumham 
        (termasuk kategori ini: Buku/Modul Ajar")""", 
        """Membuat rancangan dan karya teknologi yang tidak dipatenkan; rancangan dan karya seni monumental yang tidak terdaftar di HAKI 
        tetapi telah dipresentasikan pada forum yang teragenda - Nasional"""]

        list_10 = ["Book chapter (nasional)", 
        "Jurnal nasional", 
        "Jurnal ilmiah yang ditulis dalam Bahasa Resmi PBB namun tidak memenuhi syarat syarat sebagai jurnal ilmiah internasional",
        "Dipresentasikan secara oral dan dimuat dalam prosiding yang dipublikasikan (ber ISSN/ISBN) - Nasional",
        "Disajikan dalam bentuk poster dan dimuat dalam prosiding yang dipublikasikan - Internasional",
        "Hasil penelitian/pemikiran yang tidak disajikan dalam seminar/simposium/lokakarya, tetapi dimuat dalam prosiding - Internasional",
        "Hasil penelitian/pemikiran yang tidak disajikan dalam seminar/simposium/lokakarya, tetapi dimuat dalam prosiding - Mengedit/menyunting karya ilmiah, diterbitkan dan diedarkan secara nasional",
        "Membuat rancangan dan karya teknologi yang tidak dipatenkan; rancangan dan karya seni monumental yang tidak terdaftar di HAKI tetapi telah dipresentasikan pada forum yang teragenda - Lokal"]

        list_5 = ["Disajikan dalam bentuk poster dan dimuat dalam prosiding yang dipublikasikan - Nasional", 
        "Disajikan dalam seminar/simposium/lokakarya, tetapi tidak dimuat dalam prosiding yang dipublikasikan - Internasional (bukti sertifikat)", 
        "Hasil penelitian/pemikiran yang tidak disajikan dalam seminar/simposium/lokakarya, tetapi dimuat dalam prosiding - Nasional"]

        list_3 = ["Disajikan dalam seminar/simposium/lokakarya, tetapi tidak dimuat dalam prosiding yang dipublikasikan - Nasional (bukti sertifikat)"]

        list_2 = ["Hasil penelitian/pemikiran yang tidak disajikan dalam seminar/simposium/lokakarya, tetapi dimuat dalam prosiding - Hasil penelitian atau pemikiran atau kerjasama industri yang tidak dipublikasikan (tersimpan dalam perpustakaan) yang dilakukan secara melembaga"]
        
        list_1 = ["Hasil penelitian/pemikiran yang tidak disajikan dalam seminar/simposium/lokakarya, tetapi dimuat dalam prosiding - Hasil penelitian/pemikiran yang disajikan dalam koran/majalah populer/umum"]

        ## Max = 60:
        if (karil.category in list_60):
            max_nilai = 60

        ## Max = 40:
        elif (karil.category in list_40):
            max_nilai = 40
        
        ## Max = 30:
        elif (karil.category in list_30):
            max_nilai = 30

        ## Max = 25:
        elif (karil.category in list_25):
            max_nilai = 25
        
        ## Max = 20:
        elif (karil.category in list_20):
            max_nilai = 20
        
        ## Max = 15:
        elif (karil.category in list_15):
            max_nilai = 15

        ## Max = 10:
        elif (karil.category in list_10):
            max_nilai = 10
        
        ## Max = 5:
        elif (karil.category in list_5):
            max_nilai = 5
        
        ## Max = 3:
        elif (karil.category in list_3):
            max_nilai = 3
        
        ## Max = 2:
        elif (karil.category in list_2):
            max_nilai = 2
        
        ## Max = 1:
        elif (karil.category in list_1):
            max_nilai = 1

        else: max_nilai = 0

        worksheet1.write_number('E64', 0.1*max_nilai, merge_format_bordered)  # Nilai Max Kelengkapan dan kesesuaian unsur publikasi (10%)
        worksheet1.write_number('E65', 0.3*max_nilai, merge_format_bordered) # Nilai Max Ruang lingkup,  kedalaman pembahasan, keterbaruan (30%)
        worksheet1.write_number('E66', 0.3*max_nilai, merge_format_bordered) # Nilai Max Kecukupan dan kemutakhiran data/informasi dan metodologi (30%)
        worksheet1.write_number('E67', 0.3*max_nilai, merge_format_bordered) # Nilai Max Kelengkapan unsur dan kualitas penerbit, hasil dan manfaat (30%)
        worksheet1.write_number('E68', max_nilai, merge_format_bordered) # Nilai Max dari category (100%)

        worksheet1.merge_range('F64:G64', '', merge_format_bordered) # Nilai Review (empty)
        worksheet1.merge_range('F65:G65', '', merge_format_bordered) # Nilai Review (empty)
        worksheet1.merge_range('F66:G66', '', merge_format_bordered) # Nilai Review (empty)
        worksheet1.merge_range('F67:G67', '', merge_format_bordered) # Nilai Review (empty)
        worksheet1.merge_range('F68:G68', '', merge_format_bordered) # Nilai Total Review (empty)

        worksheet1.write('A69', 'K', merge_format_number_bordered)
        worksheet1.set_row(68, 30, merge_format)
        worksheet1.merge_range('B69:G69', 'Nilai Pengusul (pilih salah satu kriteria dibawah ini): ', merge_format_justified_category_bordered)

        worksheet1.write('A70', 'K1', merge_format_number_bordered)
        worksheet1.set_row(69, 30, merge_format)
        worksheet1.merge_range('B70:E70', 'Nilai Pengusul (penulis mandiri sekaligus koresponden) = 100%', merge_format_justified_category_bordered)
        worksheet1.merge_range('F70:G70', '', merge_format_justified_category_bordered)

        worksheet1.write('A71', 'K2', merge_format_number_bordered)
        worksheet1.set_row(70, 30, merge_format)
        worksheet1.merge_range('B71:E71', 'Nilai Pengusul (penulis pertama sekaligus koresponden) = 60%', merge_format_justified_category_bordered)
        worksheet1.merge_range('F71:G71', '', merge_format_justified_category_bordered)

        worksheet1.write('A72', 'K3', merge_format_number_bordered)
        worksheet1.set_row(71, 50, merge_format)
        worksheet1.merge_range('B72:E72', 'Nilai pengusul (penulis pertama bukan koresponden atau penulis koresponden bukan penulis pertama, jumlah penulis 2) = 50%', merge_format_justified_category_bordered)
        worksheet1.merge_range('F72:G72', '', merge_format_justified_category_bordered)

        worksheet1.write('A73', 'K4', merge_format_number_bordered)
        worksheet1.set_row(72, 50, merge_format)
        worksheet1.merge_range('B73:E73', 'Nilai pengusul (penulis pertama bukan koresponden atau penulis koresponden bukan penulis pertama, jumlah penulis > 2) = 40%', merge_format_justified_category_bordered)
        worksheet1.merge_range('F73:G73', '', merge_format_justified_category_bordered)

        worksheet1.write('A74', 'K5', merge_format_number_bordered)
        worksheet1.set_row(73, 50, merge_format)
        worksheet1.merge_range('B74:E74', 'Nilai pengusul (penulis pendamping, penulis pertama sekaligus koresponden) = 40% dibagi jumlah penulis pendamping', merge_format_justified_category_bordered)
        worksheet1.merge_range('F74:G74', '', merge_format_justified_category_bordered)

        worksheet1.write('A75', 'K6', merge_format_number_bordered)
        worksheet1.set_row(74, 50, merge_format)
        worksheet1.merge_range('B75:E75', 'Nilai pengusul (penulis pendamping, penulis pertama bukan koresponden) = 20% dibagi jumlah penulis pendamping', merge_format_justified_category_bordered)
        worksheet1.merge_range('F75:G75', '', merge_format_justified_category_bordered)

        worksheet1.set_row(75, 50, merge_format)
        worksheet1.write('B76', 'Tempat/Tanggal Review: ', merge_format_justified_category_bordered)
        worksheet1.merge_range('C76:G76', '', merge_format_justified_content)

        worksheet1.write('B77', 'Nama: ', merge_format_justified_category_bordered)
        worksheet1.merge_range('C77:G77', '', merge_format_justified_content)

        worksheet1.write('B78', 'NIP: ', merge_format_justified_category_bordered)
        worksheet1.merge_range('C78:G78', '', merge_format_justified_content)

        worksheet1.write('A79', '', merge_format_justified_category_5)
        worksheet1.write('B79', 'Unit Kerja: ', merge_format_justified_category_6)
        worksheet1.merge_range('C79:G79', '', merge_format_justified_content_2)

        logger.info("saving workbook...")
        workbook.close()

        output.seek(0)

        logger.info("serving download...")

        response = HttpResponse(output.read(), content_type = 'application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(wb_title)

        output.close()

        return response


## Handles uploading the download review form functionality
## Will read the filled data from the .xlsx file
## Will create a Review object filled with the provided data
## UPLOADED EXCEL FILE NEEDS KEY: "excel_file"
class UploadReviewForm(APIView):

    def post(self, request):
        user_data = get_user_data(request)

        excel_file = request.FILES["excel_file"]

        wb = openpyxl.load_workbook(excel_file) # openpyxl library can be read at https://openpyxl.readthedocs.io/en/stable/index.html
        ws = wb.active

        ## Check for filled score_proposer
        cell_1 = ws['F70']
        cell_2 = ws['F71']
        cell_3 = ws['F72']
        cell_4 = ws['F73']
        cell_5 = ws['F74']
        cell_6 = ws['F75']

        if cell_1.value is None:
            pass
        else:
            cp = ws['B70']
            sp = cell_1

        if cell_2.value is None:
            pass
        else:
            cp = ws['B71']
            sp = cell_2

        if cell_3.value is None:
            pass
        else:
            cp = ws['B72']
            sp = cell_3

        if cell_4.value is None:
            pass
        else:
            cp = ws['B73']
            sp = cell_4

        if cell_5.value is None:
            pass
        else:
            cp = ws['B74']
            sp = cell_5
        
        if cell_6.value is None:
            pass
        else:
            cp = ws['B75']
            sp = cell_6

        karil = KaryaIlmiah.objects.get(karil_id = ws['H1'].value)

        # Create a review based on excel data
        review_form = Review.objects.create( 
            karil_id = karil,   # This has to be a KaryaIlmiah instance
            reviewer = request.user,    # This has to be a user instance
            plagiarism_percentage = ws['C60'].value,
            linearity = ws['C61'].value,
            score_1 = Decimal(ws['F64'].value), # Use .value to get the value of the specified cells
            score_2 = Decimal(ws['F65'].value),
            score_3 = Decimal(ws['F66'].value),
            score_4 = Decimal(ws['F67'].value),
            max_1 = Decimal(ws['E64'].value),
            max_2 = Decimal(ws['E65'].value),
            max_3 = Decimal(ws['E66'].value),
            max_4 = Decimal(ws['E67'].value),
            max_total= ws['E68'].value,
            score_total = Decimal(ws['F68'].value),
            comment_1 = ws['C64'].value,
            comment_2 = ws['C65'].value,
            comment_3 = ws['C66'].value,
            comment_4 = ws['C67'].value,
            chosen_proposer = cp.value,
            score_proposer = Decimal(sp.value)
        )

        dict_obj = model_to_dict(review_form)
        serializer = ReviewSerializer(data=dict_obj)
        if serializer.is_valid():
            review = serializer.save()
            if review:
                return Response(status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Only for debugging
# class GetAllReviews(APIView):

#     def get(self, request):
#         review_list = Review.objects.all()
#         serializer = ReviewSerializer(review_list, many=True)

#         return Response(serializer.data, status=status.HTTP_200_OK)


